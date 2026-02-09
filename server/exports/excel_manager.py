"""
Gestor de Exportación Excel
GESJ - Plataforma de Gestión Educativa
"""

import os
from datetime import datetime
from typing import List, Dict, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class ExcelManager:
    """Gestor especializado para exportación a Excel"""
    
    def __init__(self):
        if not OPENPYXL_AVAILABLE:
            print("⚠️ OpenPyXL no está disponible. Instale con: pip install openpyxl")
    
    def crear_workbook_calificaciones(self, datos_calificaciones: List[Dict], 
                                    info_materia: Dict) -> str:
        """Crear workbook de Excel para calificaciones"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("OpenPyXL no está instalado")
        
        try:
            # Crear directorio de exportación
            export_dir = "exportaciones_excel"
            os.makedirs(export_dir, exist_ok=True)
            
            # Nombre del archivo
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            materia_nombre = info_materia.get('nombre', 'Materia').replace(' ', '_')
            curso = info_materia.get('curso', 'Curso').replace(' ', '')
            division = info_materia.get('division', 'A')
            
            filename = f"Calificaciones_{materia_nombre}_{curso}_{division}_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Calificaciones"
            
            # Aplicar estilos
            self._aplicar_estilos_calificaciones(ws, datos_calificaciones, info_materia)
            
            # Guardar archivo
            wb.save(filepath)
            return os.path.abspath(filepath)
            
        except Exception as e:
            print(f"Error al crear Excel de calificaciones: {e}")
            raise e
    
    def crear_workbook_promedios(self, datos_promedios: List[Dict], 
                               info_curso: Dict) -> str:
        """Crear workbook de Excel para promedios"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("OpenPyXL no está instalado")
        
        try:
            export_dir = "exportaciones_excel"
            os.makedirs(export_dir, exist_ok=True)
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            curso = info_curso.get('curso', 'Curso').replace(' ', '')
            division = info_curso.get('division', 'A')
            
            filename = f"Promedios_{curso}_{division}_{timestamp}.xlsx"
            filepath = os.path.join(export_dir, filename)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Promedios"
            
            self._aplicar_estilos_promedios(ws, datos_promedios, info_curso)
            
            wb.save(filepath)
            return os.path.abspath(filepath)
            
        except Exception as e:
            print(f"Error al crear Excel de promedios: {e}")
            raise e
    
    def _aplicar_estilos_calificaciones(self, ws, datos, info_materia):
        """Aplicar estilos específicos para calificaciones"""
        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # Título principal
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Calificaciones - {info_materia.get('nombre', 'Materia')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Información general
        row = 3
        info_data = [
            ['Materia:', info_materia.get('nombre', 'N/A')],
            ['Curso:', f"{info_materia.get('curso', 'N/A')} - División {info_materia.get('division', 'A')}"],
            ['Docente:', info_materia.get('docente', 'N/A')],
            ['Fecha de exportación:', datetime.now().strftime("%d/%m/%Y %H:%M")]
        ]
        
        for info in info_data:
            ws[f'A{row}'] = info[0]
            ws[f'A{row}'].font = Font(bold=True)
            ws[f'B{row}'] = info[1]
            row += 1
        
        # Encabezados de datos
        row += 2
        headers = ['Alumno', 'Tipo Evaluación', 'Nota', 'Fecha', 'Observaciones']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        
        # Datos de calificaciones
        row += 1
        for dato in datos:
            ws.cell(row=row, column=1, value=dato.get('alumno', '')).border = border
            ws.cell(row=row, column=2, value=dato.get('tipo_evaluacion', '')).border = border
            ws.cell(row=row, column=3, value=float(dato.get('nota', 0))).border = border
            ws.cell(row=row, column=4, value=dato.get('fecha_evaluacion', '')).border = border
            ws.cell(row=row, column=5, value=dato.get('observaciones', '')).border = border
            row += 1
        
        # Ajustar ancho de columnas
        for col in range(1, 6):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _aplicar_estilos_promedios(self, ws, datos, info_curso):
        """Aplicar estilos específicos para promedios"""
        # Similar a calificaciones pero adaptado para promedios
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
        
        # Título
        ws.merge_cells('A1:E1')
        ws['A1'] = f"Promedios - {info_curso.get('curso', 'Curso')} {info_curso.get('division', 'A')}"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # Encabezados
        headers = ['Alumno', 'Materia', 'Promedio', 'Evaluaciones', 'Estado']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos
        row = 4
        for dato in datos:
            ws.cell(row=row, column=1, value=dato.get('alumno', ''))
            ws.cell(row=row, column=2, value=dato.get('materia', ''))
            ws.cell(row=row, column=3, value=float(dato.get('promedio', 0)))
            ws.cell(row=row, column=4, value=int(dato.get('cantidad_notas', 0)))
            
            # Estado basado en promedio
            promedio = float(dato.get('promedio', 0))
            if promedio >= 9.0:
                estado = "Excelente"
            elif promedio >= 8.0:
                estado = "Muy Bueno"
            elif promedio >= 7.0:
                estado = "Bueno"
            elif promedio >= 6.0:
                estado = "Regular"
            else:
                estado = "En Riesgo"
            
            ws.cell(row=row, column=5, value=estado)
            row += 1